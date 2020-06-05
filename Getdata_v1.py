import openpyxl, cx_Oracle, json, sys
# Процедура для создания промежутка из букв
def char_range(c1, c2):
    for c in range(ord(c1), ord(c2)+1):
        yield chr(c)
#Загоняем в вб информацию о рабочем екселе
myconnection =cx_Oracle.connect('SHOES', 'SHOES', '//localhost:1521/xe', encoding='utf8', nencoding='utf8')
mycursor = myconnection.cursor()
wb = openpyxl.load_workbook('test.xlsx')
# печатаем список листов
sheets = wb.sheetnames
#Выбираем активную страницу в екселе
sheet = wb['Nts']
string = [[],[],[],[],[]]
size_range_list = [[],[],[]]
k = 1
s = 1
#Цикл проходящий от начала до конца заполненных ячеек 
while (str(sheet.cell(row = k+1, column = 6).value)!='None'):
    k+=1
    s+=1
    z = 0
    for i in char_range('A','E'):
        #Присваиваем zому элементу массива string значения рабочей ячейки (i по буквам, str(s) по строкам)
        string[z].append(str(sheet[i+str(s)].value))    
        z+=1
        # аполняем массив каждым уникальным значением Названия обуви
    size_range_list[0].append(str(sheet.cell(row = k, column = 6).value))
    size = []
    quantity = []
    #цикл для заполнения массивов, хранящих в себе рамзеры обуви и их количество в упаковке 
    while(str(sheet.cell(row = k, column = 6).value) == str(sheet.cell(row = k+1, column = 6).value)):
        size.append(str(sheet.cell(row = k, column = 7).value))   
        quantity.append(str(sheet.cell(row = k, column = 8).value))
        k+=1
    size.append(str(sheet.cell(row = k, column = 7).value))   
    quantity.append(str(sheet.cell(row = k, column = 8).value))
    #Заполнения информации о размере и кол-ве в основной масив
    size_range_list[1].append(size)
    size_range_list[2].append(quantity)
print(string)
print(size_range_list)
myconnection =cx_Oracle.connect('shoes', 'shoes', 'ORCL', encoding='utf8', nencoding='utf8')
mycursor = myconnection.cursor()
#shoes_id IMG PRICE1 category_name sub_category_name shoes_name size quantity
for i in range(len(string[0])):
    #print("add_new_shoes_full("+str(string[0][i])+",'"+str(size_range_list[0][i])+"','"+str(string[1][i])+"',"+ str(string[2][i])+",'"+ str(string[3][i])+"','"+str(string[4][i])+"')")
    mycursor.callproc('add_new_shoes_full',[string[0][i],size_range_list[0][i],string[1][i],string[2][i],string[3][i],string[4][i]])
    print(str(i)+'i')
    k = 0
for i in range(len(size_range_list[0])):
    for j in range(len(size_range_list[1][i])):
        print(str(k)+'l')
        k+=1
        #print("add_new_size_range('"+str(size_range_list[0][i])+"','"+str(size_range_list[1][i][j])+"',"+str(size_range_list[2][i][j])+")")
        mycursor.callproc('size_range_insert_update',[size_range_list[0][i],size_range_list[1][i][j],size_range_list[2][i][j]])
myconnection.commit()
myconnection.close()
